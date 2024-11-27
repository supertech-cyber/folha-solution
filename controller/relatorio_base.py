import os
from openpyxl import load_workbook
import re
from controller.regex_ import get_competencia_provisao, get_cnpj


class RelatorioBase:

    def __init__(self):
        # self.folder_default = 'relatórios'
        self.relatorios_folha = []

    def _read_path_files(self, folder_default):

        files = [os.path.join(folder_default, file) for file in os.listdir(
            folder_default) if os.path.splitext(file)[1] == '.xlsx']
        return files

    def _find_all_cnpjs(self, relatorios, sheet):

        conjunto_cnpjs = set()
        # Path to the Excel file
        excel_file_path = relatorios

        # Load the Excel file
        workbook = load_workbook(excel_file_path)

        # Assuming the data is in the first sheet
        sheet = workbook[sheet]

        # Regular expression to match CNPJs
        cnpj_pattern = r'\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b'

        # Iterate through all cells in the sheet
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if isinstance(cell_value, str):
                    # Find all CNPJs in the cell value
                    cnpjs = re.findall(cnpj_pattern, cell_value)
                    if cnpjs:
                        # Print all found CNPJs
                        for cnpj in cnpjs:
                            conjunto_cnpjs.add(cnpj)

        workbook.close()

        return conjunto_cnpjs

    def _find_all_cnpjs_plan_folha(self, relatorios, sheet):

        conjunto_cnpjs = set()
        # Path to the Excel file
        excel_file_path = relatorios

        # Load the Excel file
        workbook = load_workbook(excel_file_path)

        # Assuming the data is in the first sheet
        # sheet = workbook[sheet]
        sheet = workbook.active

        # Regular expression to match CNPJs
        cnpj_pattern = r'\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b'

        # Iterate through all cells in the sheet
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if isinstance(cell_value, str):
                    # Find all CNPJs in the cell value
                    cnpjs = re.findall(cnpj_pattern, cell_value)
                    if cnpjs:
                        # Print all found CNPJs
                        for cnpj in cnpjs:
                            conjunto_cnpjs.add(cnpj)

        workbook.close()

        return conjunto_cnpjs

    def _find_banco_dados_folha(self, cnpjs_relatorios):
        lista_bancos_dados = [os.path.join('banco de dados', file) for file in os.listdir(
            'banco de dados') if os.path.splitext(file)[1] == '.xlsx']

        for banco_dados in lista_bancos_dados:
            cnpjs_banco = self._find_all_cnpjs(banco_dados, 'empresas')
            intersection_lines = cnpjs_relatorios.intersection(cnpjs_banco)
            if intersection_lines:
                return banco_dados

        return False

    def _find_word(self, word_to_find, file_xlsx):

        found_word = False
        # Load the Excel file
        workbook = load_workbook(file_xlsx)

        # Assuming your data is in the first sheet
        sheet = workbook.active

        # Iterate through all cells in the sheet
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if isinstance(cell_value, str) and word_to_find in cell_value:
                    found_word = True
                    break

        workbook.close()

        return found_word

    def find_relatorios_folha(self, word_to_find):

        files = self._read_path_files('relatórios')

        for relatorios in files:
            if self._find_word(word_to_find, relatorios):
                cnpjs_relatorio = self._find_all_cnpjs_plan_folha(
                    relatorios, 'Page 1')

                address_db = self._find_banco_dados_folha(
                    cnpjs_relatorios=cnpjs_relatorio)

                if address_db:

                    self.relatorios_folha.append(
                        Folha(address_folha=relatorios, address_db=address_db))

    def find_banco_dados(self):

        for relatorios in self.relatorios_folha:
            pass


class Folha:

    def __init__(self, address_folha, address_db):
        self.address_folha = address_folha
        self.address_db = address_db
        self.log = None
        self.filename_out = ''


if __name__ == '__main__':

    manager = RelatorioBase()
    manager.find_relatorios_folha()
    for file in manager.relatorios_folha:
        print(f'address_folha {file.address_folha} | address_db {
              file.address_db}')
