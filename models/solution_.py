from openpyxl import load_workbook

from openpyxl.styles.colors import WHITE, RGB

__old_rgb_set__ = RGB.__set__


def __rgb_set_fixed__(self, instance, value):
    try:
        __old_rgb_set__(self, instance, value)
    except ValueError as e:
        if e.args[0] == 'Colors must be aRGB hex values':
            __old_rgb_set__(self, instance, WHITE)  # Change default color here


RGB.__set__ = __rgb_set_fixed__


class Solution:

    def __init__(self, address_db):
        self.address_db = address_db
        self.planilhas = {}

    def _read_sheets(self):

        for sheet_name in self.workbook.sheetnames:

            sheet = self.workbook[sheet_name]

            # Obtém os cabeçalhos (primeira linha)
            cabecalho = [cell.value for cell in sheet[1]]

            dados_dict = {}

            for row in sheet.iter_rows(min_row=2, values_only=True):

                linha_dict = dict(zip(cabecalho, row))
                dados_dict[str(row[0])] = linha_dict

            self.planilhas[sheet_name] = dados_dict

    def read_plan_bd(self):

        self.workbook = load_workbook(filename=self.address_db)

        self._read_sheets()


if __name__ == '__main__':

    output_solution = Solution('banco de dados\\banco_dados_TAJ.xlsx')
    output_solution.read_plan_bd()
    print(output_solution.planilhas['custo'])
